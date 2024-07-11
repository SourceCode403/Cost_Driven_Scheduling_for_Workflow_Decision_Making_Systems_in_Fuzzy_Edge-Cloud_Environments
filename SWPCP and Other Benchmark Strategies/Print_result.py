import csv
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from Fuzzy_operations import fuzzy_fitness


def reset_col(filename):  # 自动列宽
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen + 4  # 也就是列宽为最大长度*1.2+4 可以自己调整
        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(filename)


def csv_to_xlsx_pd(filename):  # csv转xlsx
    pd_csv = pd.read_csv(filename + '.csv', header=None, encoding='utf-8')
    if pd_csv.shape[0] < 20:
        pd_csv.index = list(range(1, pd_csv.shape[0]-2)) + ["avg.", "wor.", "opt."]
    else:
        pd_csv.index = list(range(1, pd_csv.shape[0]+1))
    pd_csv.to_excel(filename + '.xlsx', header=None, sheet_name='data')

    reset_col(filename + '.xlsx')
    if os.path.exists(filename + '.csv'):  # 如果文件存在
        os.remove(filename + '.csv')  # 删除csv文件


def Print_Result(g_best, iteration_cost, iteration_time, result, flag, num):
    runcost = [round(i, 3) for i in g_best.run_cost]
    makespan = [round(i, 3) for i in g_best.makespan]
    if flag:
        runtime = [[round(j, 3) for j in i] for i in g_best.run_time]
        # runtime = []
        # for i in g_best.run_time:
        #    runtime.append([round(j, 3) for j in i])
        print("Generation: {}".format(num), end="\n ")
        iteration_cost[num] += runcost + [' ']
        iteration_time[num] += [g_best.meet_deadline]
        iteration_time[num] += makespan + [' ']
        for i in runtime:
            iteration_time[num] += i + [' ']
    else:
        runtime = [round(j, 3) for i in g_best.run_time for j in i]
        print("The {:d}th Execution Result: ".format(num), end="\n ")
        result.append([tuple(runcost)] + [round(fuzzy_fitness(runcost), 3)] + [tuple(makespan)] +
                      [g_best.meet_deadline == len(g_best.run_time)] + [g_best.meet_deadline] + runtime)
    print("g_best.TS: ", [round(i, 3) for i in g_best.TS], '\n',
          "g_best.SA: ", [round(i, 3) for i in g_best.SA], '\n',
          "Com_cost: ", [round(i, 3) for i in g_best.com_cost], '  ',
          "Tran_cost: ", [round(i, 3) for i in g_best.tran_cost], '\n',
          "Run_cost: ", runcost, '  ',
          "Run_time: ", runtime, '\n',
          "Makespan: ", [round(i, 3) for i in g_best.makespan], '  ',
          "Meet_Deadline: ", g_best.meet_deadline)


def Output_File(Output_path, algorithm, base, iteration_cost, iteration_time, result):
    with open(Output_path + algorithm + '_' + str(base) + '_iteration_cost.csv', 'w', newline='') as file:
        writer = csv.writer(file, dialect='excel')
        writer.writerows(iteration_cost[1:])
    with open(Output_path + algorithm + '_' + str(base) + '_iteration_time.csv', 'w', newline='') as file:
        writer = csv.writer(file, dialect='excel')
        writer.writerows(iteration_time[1:])
    with open(Output_path + algorithm + '_' + str(base) + '_result.csv', 'w', newline='') as file:
        writer = csv.writer(file, dialect='excel')
        writer.writerows(result)
    filename = ['iteration_cost', 'iteration_time', 'result']
    for i in range(3):
        csv_to_xlsx_pd(Output_path + algorithm + '_' + str(base) + '_' + filename[i])
